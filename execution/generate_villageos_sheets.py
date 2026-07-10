import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_villageos_sheet():
    wb = openpyxl.Workbook()
    
    # Clean default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    tabs = [
        "00_Settings",
        "01_Dashboard",
        "02_Village_Data",
        "03_Resources",
        "04_Hero",
        "05_Culture",
        "06_Build_Planner",
        "07_Raid_Planner",
        "08_Map_Planner",
        "09_Daily_Log",
        "10_AI_Analysis",
        "Calculations_Lookup" # Core game data database sheet
    ]
    
    color_dark_primary = "1E293B"  # Slate 800
    color_light_gray = "F8FAFC"    # Slate 50
    color_border_slate = "E2E8F0"  # Slate 200
    color_green = "10B981"         # Green 500
    color_blue = "3B82F6"          # Blue 500
    color_orange = "F59E0B"        # Amber 500
    
    font_family = "Inter"
    font_title = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    font_section = Font(name=font_family, size=11, bold=True, color="000000")
    font_bold = Font(name=font_family, size=10, bold=True)
    font_regular = Font(name=font_family, size=10)
    font_kpi_val = Font(name=font_family, size=14, bold=True, color="1E293B")
    
    fill_header = PatternFill(start_color=color_dark_primary, end_color=color_dark_primary, fill_type="solid")
    fill_card = PatternFill(start_color=color_light_gray, end_color=color_light_gray, fill_type="solid")
    fill_row_alt = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    
    thin_side = Side(border_style="thin", color=color_border_slate)
    border_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    for tab_name in tabs:
        ws = wb.create_sheet(title=tab_name)
        
        # Hide the database lookup sheet to keep the UI clean on mobile
        if tab_name == "Calculations_Lookup":
            ws.sheet_state = 'hidden'
            
        ws.views.sheetView[0].showGridLines = False
        
        # Header banner
        ws.merge_cells("A1:M1")
        ws["A1"] = f"  VillageOS  |  {tab_name[3:].replace('_', ' ')}"
        ws["A1"].font = font_title
        ws["A1"].fill = fill_header
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 40
        
        ws.row_dimensions[2].height = 15
        
        if tab_name == "00_Settings":
            ws["A3"] = "Key"
            ws["B3"] = "Value"
            ws["A3"].font = font_bold
            ws["B3"].font = font_bold
            ws["A3"].border = border_thin
            ws["B3"].border = border_thin
            
            settings = [
                ("Account Name", "Swayam"),
                ("Tribe", "Gauls"),
                ("Server Speed", 3),
                ("Gold Strategy", "Light Gold"),
                ("Timezone", "GMT+5:30"),
                ("Server Start Date", "2026-07-01"),
                ("Current Village", "Capital 01"),
                ("Capital Coordinates", "(0|0)"),
                ("Target Expansion Slot", 2)
            ]
            for idx, (k, v) in enumerate(settings, start=4):
                ws.row_dimensions[idx].height = 22
                ws[f"A{idx}"] = k
                ws[f"B{idx}"] = v
                ws[f"A{idx}"].font = font_regular
                ws[f"B{idx}"].font = font_regular
                ws[f"A{idx}"].border = border_thin
                ws[f"B{idx}"].border = border_thin
                
        elif tab_name == "01_Dashboard":
            ws.merge_cells("B4:D4")
            ws["B4"] = "Active Profile"
            ws["B4"].font = font_bold
            ws["B4"].fill = fill_card
            ws["B4"].alignment = align_center
            ws["B4"].border = border_thin
            
            ws.merge_cells("B5:D5")
            ws["B5"] = "='00_Settings'!B4 & \" (x\" & '00_Settings'!B6 & \")\""
            ws["B5"].font = font_kpi_val
            ws["B5"].fill = fill_card
            ws["B5"].alignment = align_center
            ws["B5"].border = border_thin
            
            ws.merge_cells("F4:H4")
            ws["F4"] = "Settlement ETA"
            ws["F4"].font = font_bold
            ws["F4"].fill = fill_card
            ws["F4"].alignment = align_center
            ws["F4"].border = border_thin
            
            ws.merge_cells("F5:H5")
            ws["F5"] = "='05_Culture'!B8"
            ws["F5"].font = font_kpi_val
            ws["F5"].fill = fill_card
            ws["F5"].alignment = align_center
            ws["F5"].border = border_thin
            
            resource_headers = [
                ("Wood Status", "='03_Resources'!B4 & \" / \" & '03_Resources'!D4", "='03_Resources'!C4 & \"/hr\"", "B", "C"),
                ("Clay Status", "='03_Resources'!E4 & \" / \" & '03_Resources'!G4", "='03_Resources'!F4 & \"/hr\"", "E", "F"),
                ("Iron Status", "='03_Resources'!H4 & \" / \" & '03_Resources'!J4", "='03_Resources'!I4 & \"/hr\"", "H", "I"),
                ("Crop Status", "='03_Resources'!K4 & \" / \" & '03_Resources'!M4", "='03_Resources'!L4 & \"/hr\"", "K", "L")
            ]
            
            row_idx = 7
            for title, formula_val, formula_prod, col1, col2 in resource_headers:
                ws.row_dimensions[row_idx].height = 20
                ws.row_dimensions[row_idx+1].height = 22
                
                ws.merge_cells(f"{col1}{row_idx}:{col2}{row_idx}")
                ws[f"{col1}{row_idx}"] = title
                ws[f"{col1}{row_idx}"].font = font_bold
                ws[f"{col1}{row_idx}"].fill = fill_card
                ws[f"{col1}{row_idx}"].alignment = align_center
                ws[f"{col1}{row_idx}"].border = border_thin
                
                ws.merge_cells(f"{col1}{row_idx+1}:{col2}{row_idx+1}")
                ws[f"{col1}{row_idx+1}"] = formula_val
                ws[f"{col1}{row_idx+1}"].font = font_kpi_val
                ws[f"{col1}{row_idx+1}"].fill = fill_card
                ws[f"{col1}{row_idx+1}"].alignment = align_center
                ws[f"{col1}{row_idx+1}"].border = border_thin
                
                ws.merge_cells(f"{col1}{row_idx+2}:{col2}{row_idx+2}")
                ws[f"{col1}{row_idx+2}"] = formula_prod
                ws[f"{col1}{row_idx+2}"].font = font_regular
                ws[f"{col1}{row_idx+2}"].fill = fill_card
                ws[f"{col1}{row_idx+2}"].alignment = align_center
                ws[f"{col1}{row_idx+2}"].border = border_thin
                
                row_idx += 4
                
        elif tab_name == "02_Village_Data":
            headers = ["Village ID", "Category", "Slot ID", "Building Name", "Current Level", "Target Level", "Priority"]
            for col_idx, text in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=col_idx, value=text)
                cell.font = font_bold
                cell.border = border_thin
            ws.row_dimensions[3].height = 22
            
            mock_buildings = [
                (1, "d2", 1, "Woodcutter", 4, 10, 5),
                (1, "d2", 2, "Clay Pit", 5, 10, 5),
                (1, "d2", 3, "Iron Mine", 4, 10, 4),
                (1, "d2", 4, "Cropland", 4, 10, 5),
                (1, "d1", 19, "Main Building", 5, 10, 3),
                (1, "d1", 20, "Warehouse", 3, 5, 4),
                (1, "d1", 21, "Granary", 3, 5, 4),
            ]
            for row_idx, data in enumerate(mock_buildings, start=4):
                ws.row_dimensions[row_idx].height = 22
                is_alt = (row_idx % 2 == 0)
                for col_idx, val in enumerate(data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = font_regular
                    cell.border = border_thin
                    if is_alt:
                        cell.fill = fill_row_alt
                        
        elif tab_name == "03_Resources":
            headers = ["Village ID", "Wood Current", "Wood Prod", "Wood Cap", 
                       "Clay Current", "Clay Prod", "Clay Cap", 
                       "Iron Current", "Iron Prod", "Iron Cap", 
                       "Crop Current", "Crop Prod", "Crop Cap"]
            for col_idx, text in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=col_idx, value=text)
                cell.font = font_bold
                cell.border = border_thin
            ws.row_dimensions[3].height = 22
            
            ws.row_dimensions[4].height = 22
            ws["A4"] = 1
            ws["B4"] = 12000
            ws["C4"] = 420
            ws["D4"] = 20000
            ws["E4"] = 14500
            ws["F4"] = 480
            ws["G4"] = 20000
            ws["H4"] = 6200
            ws["I4"] = 320
            ws["J4"] = 20000
            ws["K4"] = 1500
            ws["L4"] = 120
            ws["M4"] = 20000
            
            for col_letter in [get_column_letter(i) for i in range(1, 14)]:
                ws[f"{col_letter}4"].font = font_regular
                ws[f"{col_letter}4"].border = border_thin

        elif tab_name == "04_Hero":
            ws["A3"] = "Attribute"
            ws["B3"] = "Value"
            ws["A3"].font = font_bold
            ws["B3"].font = font_bold
            ws["A3"].border = border_thin
            ws["B3"].border = border_thin

            hero_stats = [
                ("Hero Name", "Jack_Slayer"),
                ("Level", 6),
                ("Current XP", 1850),
                ("Next Level XP", 2500),
                ("Health %", 85),
                ("Fighting Strength", 480),
                ("Resource Production Bonus %", 24),
                ("Active Weapon", "Sword of the Gaul"),
                ("Active Armor", "Plated Chestplate"),
                ("Adventures Run", 12),
                ("Current Status", "Idle")
            ]
            for idx, (k, v) in enumerate(hero_stats, start=4):
                ws.row_dimensions[idx].height = 22
                ws[f"A{idx}"] = k
                ws[f"B{idx}"] = v
                ws[f"A{idx}"].font = font_regular
                ws[f"B{idx}"].font = font_regular
                ws[f"A{idx}"].border = border_thin
                ws[f"B{idx}"].border = border_thin

        elif tab_name == "05_Culture":
            ws["A3"] = "Metric"
            ws["B3"] = "Value"
            ws["A3"].font = font_bold
            ws["B3"].font = font_bold
            ws["A3"].border = border_thin
            ws["B3"].border = border_thin
            
            ws["A4"] = "Current CP"
            ws["B4"] = 423
            ws["A5"] = "CP / Day"
            ws["B5"] = 120
            ws["A6"] = "Next Slot CP Goal"
            ws["B6"] = 2000
            ws["A7"] = "Days Until Expansion"
            ws["B7"] = "=(B6-B4)/B5"
            ws["A8"] = "Settlement Date ETA"
            ws["B8"] = "=TODAY()+B7"
            
            for r in range(4, 9):
                ws.row_dimensions[r].height = 22
                ws[f"A{r}"].font = font_regular
                ws[f"B{r}"].font = font_regular
                ws[f"A{r}"].border = border_thin
                ws[f"B{r}"].border = border_thin
                
        elif tab_name == "06_Build_Planner":
            headers = ["Status", "Village", "Building", "Level", "Time Required (s)", "Wood Cost", "Clay Cost", "Iron Cost", "Crop Cost"]
            for col_idx, text in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=col_idx, value=text)
                cell.font = font_bold
                cell.border = border_thin
            ws.row_dimensions[3].height = 22
            
            ws.row_dimensions[4].height = 22
            ws["A4"] = "NOW"
            ws["B4"] = "Capital 01"
            ws["C4"] = "Clay Pit"
            ws["D4"] = 6
            # Duration Formula VLOOKUP to Calculations_Lookup
            ws["E4"] = "=ROUND((VLOOKUP(C4, Calculations_Lookup!$A$4:$H$20, 8, FALSE) * 1.5) / '00_Settings'!B6 * (0.96 ^ VLOOKUP(\"Main Building\", '02_Village_Data'!$D$4:$F$10, 2, FALSE)))"
            # Resource Costs Formulas VLOOKUP to Calculations_Lookup scaled geometrically by 1.28^(L-1)
            ws["F4"] = "=ROUND(VLOOKUP(C4, Calculations_Lookup!$A$4:$H$20, 2, FALSE) * (1.28 ^ (D4 - 1)))"
            ws["G4"] = "=ROUND(VLOOKUP(C4, Calculations_Lookup!$A$4:$H$20, 3, FALSE) * (1.28 ^ (D4 - 1)))"
            ws["H4"] = "=ROUND(VLOOKUP(C4, Calculations_Lookup!$A$4:$H$20, 4, FALSE) * (1.28 ^ (D4 - 1)))"
            ws["I4"] = "=ROUND(VLOOKUP(C4, Calculations_Lookup!$A$4:$H$20, 5, FALSE) * (1.28 ^ (D4 - 1)))"
            
            for col_letter in [get_column_letter(i) for i in range(1, 10)]:
                ws[f"{col_letter}4"].font = font_regular
                ws[f"{col_letter}4"].border = border_thin
                
        elif tab_name == "07_Raid_Planner":
            headers = ["Oasis Coordinates", "Distance (Cells)", "Animals Present", "Loot Capacity", "HP Loss %", "ROI Score"]
            for col_idx, text in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=col_idx, value=text)
                cell.font = font_bold
                cell.border = border_thin
            ws.row_dimensions[3].height = 22

            raid_targets = [
                ("(-4|12)", 3.6, "None", 800, 2, "=D4/(B4+1)"),  
                ("(2|-8)", 8.2, "Spiders", 1500, 15, "=D5/(B5+1)"),
                ("(-1|1)", 1.4, "Bats", 400, 5, "=D6/(B6+1)")
            ]
            for row_idx, data in enumerate(raid_targets, start=4):
                ws.row_dimensions[row_idx].height = 22
                is_alt = (row_idx % 2 == 0)
                for col_idx, val in enumerate(data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = font_regular
                    cell.border = border_thin
                    if is_alt:
                        cell.fill = fill_row_alt

        elif tab_name == "08_Map_Planner":
            headers = ["Coordinates", "Distance", "Cropper Type", "Oasis Bonus %", "Owner", "Priority Score"]
            for col_idx, text in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=col_idx, value=text)
                cell.font = font_bold
                cell.border = border_thin
            ws.row_dimensions[3].height = 22

            map_data = [
                ("(12|-4)", 12.6, "15c", 75, "Unoccupied", "=D4/(B4+1)"),
                ("(-10|2)", 10.2, "9c", 50, "Unoccupied", "=D5/(B5+1)"),
                ("(1|15)", 15.0, "15c", 100, "Unoccupied", "=D6/(B6+1)")
            ]
            for row_idx, data in enumerate(map_data, start=4):
                ws.row_dimensions[row_idx].height = 22
                is_alt = (row_idx % 2 == 0)
                for col_idx, val in enumerate(data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = font_regular
                    cell.border = border_thin
                    if is_alt:
                        cell.fill = fill_row_alt

        elif tab_name == "09_Daily_Log":
            headers = ["Date", "Population", "Hero Level", "Wood Prod", "Clay Prod", "Iron Prod", "Crop Prod", "Total CP", "Rank"]
            for col_idx, text in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=col_idx, value=text)
                cell.font = font_bold
                cell.border = border_thin
            ws.row_dimensions[3].height = 22

            log_data = [
                ("2026-07-01", 100, 1, 100, 100, 80, 20, 40, 2400),
                ("2026-07-02", 112, 2, 140, 140, 110, 25, 80, 2250),
                ("2026-07-03", 124, 2, 190, 195, 150, 30, 130, 2100)
            ]
            for row_idx, data in enumerate(log_data, start=4):
                ws.row_dimensions[row_idx].height = 22
                is_alt = (row_idx % 2 == 0)
                for col_idx, val in enumerate(data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = font_regular
                    cell.border = border_thin
                    if is_alt:
                        cell.fill = fill_row_alt

        elif tab_name == "10_AI_Analysis":
            ws["A3"] = "Today's Recommendations"
            ws["A3"].font = font_section
            ws.row_dimensions[3].height = 22
            
            headers = ["Priority", "Recommended Action", "Reason", "ETA / Wait Time", "Benefit Metric"]
            for col_idx, text in enumerate(headers, start=1):
                cell = ws.cell(row=4, column=col_idx, value=text)
                cell.font = font_bold
                cell.border = border_thin
            ws.row_dimensions[4].height = 22

            recs = [
                ("★★★★★", "Upgrade Clay Pit to 6", "Highest ROI", "Ready in 42m", "+48 Clay/hr"),
                ("★★★★★", "Run Adventure", "Hero Idle & Ready", "Immediate", "Expected XP & Items"),
                ("★★★★☆", "Build Marketplace Lvl 1", "Unlock Trade Options", "12m 40s", "Enables trade routes"),
                ("★★★★☆", "Attack Oasis (-4|12)", "Highest Raid ROI", "35m travel", "Expected 800 resources")
            ]
            for row_idx, data in enumerate(recs, start=5):
                ws.row_dimensions[row_idx].height = 22
                is_alt = (row_idx % 2 == 0)
                for col_idx, val in enumerate(data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = font_regular
                    cell.border = border_thin
                    if is_alt:
                        cell.fill = fill_row_alt
                        
        elif tab_name == "Calculations_Lookup":
            # Building lookup database
            ws["A3"] = "Building Name"
            ws["B3"] = "Base Wood"
            ws["C3"] = "Base Clay"
            ws["D3"] = "Base Iron"
            ws["E3"] = "Base Crop"
            ws["F3"] = "Base Pop"
            ws["G3"] = "Base CP"
            ws["H3"] = "Base Time (s)"
            
            for col in range(1, 9):
                ws.cell(row=3, column=col).font = font_bold
                ws.cell(row=3, column=col).border = border_thin
                
            building_db = [
                ("Woodcutter", 40, 100, 50, 60, 2, 1, 300),
                ("Clay Pit", 120, 40, 80, 80, 2, 1, 300),
                ("Iron Mine", 100, 80, 40, 60, 2, 1, 360),
                ("Cropland", 70, 90, 70, 20, 0, 1, 260),
                ("Main Building", 70, 40, 60, 20, 2, 2, 330),
                ("Warehouse", 130, 160, 90, 40, 1, 1, 380),
                ("Granary", 80, 100, 70, 120, 1, 1, 310),
                ("Academy", 220, 160, 90, 40, 4, 4, 450),
                ("Barracks", 70, 120, 100, 80, 4, 1, 320)
            ]
            for row_idx, data in enumerate(building_db, start=4):
                ws.row_dimensions[row_idx].height = 22
                for col_idx, val in enumerate(data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = font_regular
                    cell.border = border_thin
            
        # Optimize column widths
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                if val.startswith("="):
                    max_len = max(max_len, 14)
                else:
                    max_len = max(max_len, len(val))
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    target_path = r"c:\Users\swaya\OneDrive\Desktop\Master_AG\VillageOS\sheets\VillageOS_v1.xlsx"
    os.makedirs(os.path.dirname(target_path), exist_ok=True)
    wb.save(target_path)
    print("Database calculations lookup sheet integrated and linked successfully.")

if __name__ == "__main__":
    create_villageos_sheet()
